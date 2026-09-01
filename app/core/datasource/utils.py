"""
数据源操作共享工具方法
供任务执行和内置工具复用，避免代码重复
"""

import json
import re
import io
from datetime import datetime, date, time as dt_time
from typing import Any, List


def quote_ident(ident: str) -> str:
    """
    SQL标识符加反引号，仅含特殊字符时加引号

    Args:
        ident: 标识符名称

    Returns:
        str: 处理后的标识符
    """
    return f"`{ident}`" if not re.fullmatch(r'[A-Za-z_][A-Za-z0-9_]*', ident) else ident


def normalize_row_value(value: Any) -> Any:
    """
    将数据库行中的非JSON可序列化值（datetime/date/time/LOB/bytes等）转为字符串

    Args:
        value: 数据库查询结果中的原始值

    Returns:
        Any: 可JSON序列化的值
    """
    if value is None:
        return None
    if isinstance(value, dict):
        return {k: normalize_row_value(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [normalize_row_value(v) for v in value]
    if isinstance(value, (datetime, date, dt_time)):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    # 处理set类型
    if isinstance(value, set):
        return str(value)
    # 处理bytes类型（BLOB、BYTEA等二进制字段）
    if isinstance(value, bytes):
        try:
            return value.decode('utf-8')
        except UnicodeDecodeError:
            return '0x' + value.hex()
    # 处理Oracle LOB对象（通过类名判断，避免强依赖oracledb）
    class_name = value.__class__.__name__
    if class_name == 'LOB':
        try:
            return value.read()
        except Exception:
            try:
                return value.getvalue()
            except Exception:
                return str(value)
    # 处理其他可读取对象
    if hasattr(value, 'read') and callable(value.read):
        try:
            return value.read()
        except Exception:
            pass
    return value


def normalize_rows(rows: list) -> list:
    """
    批量转换查询结果行，保证可JSON序列化

    Args:
        rows: 原始行数据列表

    Returns:
        list: 规范化后的行数据列表
    """
    return [normalize_row_value(row) for row in rows]


def format_data(data: list, export_format: str, cancel_check_fn=None) -> bytes:
    """
    根据导出格式转换数据，返回 bytes 类型内容

    Args:
        data: 行数据列表
        export_format: 导出格式 (json/markdown/excel)
        cancel_check_fn: 可选的取消检查函数，签名 () -> bool，返回 True 表示任务已取消，
                        在生成过程中会被周期性调用以支持中断

    Returns:
        bytes: 格式化后的二进制内容

    Raises:
        InterruptedError: 如果 cancel_check_fn 返回 True（表示任务已取消）
    """
    from app.constants.ontology_constants import OntologyExportFormat

    if export_format == OntologyExportFormat.EXCEL:
        return _to_excel_bytes(data, cancel_check_fn)

    if export_format == OntologyExportFormat.JSON:
        content = json.dumps(data, ensure_ascii=False, indent=2)
        return content.encode('utf-8')

    if export_format == OntologyExportFormat.MARKDOWN:
        if data:
            headers = list(data[0].keys())
            lines = [
                '| ' + ' | '.join(headers) + ' |',
                '| ' + ' | '.join(['---'] * len(headers)) + ' |',
            ]
            for i, row in enumerate(data):
                if cancel_check_fn and cancel_check_fn():
                    raise InterruptedError("任务已取消")
                lines.append('| ' + ' | '.join(str(row.get(h, '')) for h in headers) + ' |')
            return '\n'.join(lines).encode('utf-8')
        return '（无数据）'.encode('utf-8')

    # 默认JSON格式
    content = json.dumps(data, ensure_ascii=False, indent=2)
    return content.encode('utf-8')


def format_data_to_text(data: list, export_format: str) -> str:
    """
    根据导出格式转换数据，返回字符串内容（用于工具等直接返回文本的场景）

    Args:
        data: 行数据列表
        export_format: 导出格式 (json/markdown)

    Returns:
        str: 格式化后的文本内容
    """
    from app.constants.ontology_constants import OntologyExportFormat

    if export_format == OntologyExportFormat.JSON:
        return json.dumps(data, ensure_ascii=False, indent=2)

    if export_format == OntologyExportFormat.MARKDOWN:
        if data:
            headers = list(data[0].keys())
            lines = [
                '| ' + ' | '.join(headers) + ' |',
                '| ' + ' | '.join(['---'] * len(headers)) + ' |',
            ]
            for row in data:
                lines.append('| ' + ' | '.join(str(row.get(h, '')) for h in headers) + ' |')
            return '\n'.join(lines)
        return '（无数据）'

    # 默认JSON格式
    return json.dumps(data, ensure_ascii=False, indent=2)


def _to_excel_bytes(data: list, cancel_check_fn=None) -> bytes:
    """将数据列表转换为Excel二进制内容

    Args:
        data: 行数据列表
        cancel_check_fn: 可选的取消检查函数 () -> bool
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = '数据抽取结果'

    if not data:
        ws['A1'] = '（无数据）'
        ws['A1'].font = Font(bold=True)
    else:
        headers = list(data[0].keys())
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, row_data in enumerate(data, 2):
            if cancel_check_fn and cancel_check_fn():
                raise InterruptedError("任务已取消")
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                cell.alignment = Alignment(vertical='center')

        # 自适应列宽
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_data in data:
                cell_val = str(row_data.get(header, ''))
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    # 保存文件前再检查一次
    if cancel_check_fn and cancel_check_fn():
        raise InterruptedError("任务已取消")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()