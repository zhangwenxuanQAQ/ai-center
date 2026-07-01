"""
文件工具类
提供文件类型判断、缩略图生成、音频转换等功能
"""

import re
import base64
import threading
import sys
import os
import tempfile
import logging
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, List, Optional

from app.constants.knowledgebase_document_constants import FileType
from app.core.utils.resource_utils import get_file_icon_path

logger = logging.getLogger(__name__)

FFMPEG_COMMON_LOCATIONS = [
    r"C:\Users\Admin\scoop\shims\ffmpeg.exe",
    r"C:\Program Files\ffmpeg\bin\ffmpeg.exe",
    r"C:\ffmpeg\bin\ffmpeg.exe",
    r"D:\ffmpeg\bin\ffmpeg.exe",
    r"E:\ffmpeg\bin\ffmpeg.exe",
    r"C:\Program Files (x86)\ffmpeg\bin\ffmpeg.exe",
    r"D:\Program Files\ffmpeg\bin\ffmpeg.exe",
    r"D:\anaconda\Library\bin\ffmpeg.exe",
    r"C:\Users\Admin\AppData\Local\Programs\ffmpeg\bin\ffmpeg.exe",
]

LOCK_KEY_pdfplumber = "global_shared_lock_pdfplumber"
if LOCK_KEY_pdfplumber not in sys.modules:
    sys.modules[LOCK_KEY_pdfplumber] = threading.Lock()

IMG_BASE64_PREFIX = "data:image/png;base64,"

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    pdfplumber = None
    PDFPLUMBER_AVAILABLE = False

try:
    from PIL import Image, ImageDraw
    PIL_AVAILABLE = True
except ImportError:
    Image = None
    ImageDraw = None
    PIL_AVAILABLE = False


def filename_type(filename):
    """
    根据文件扩展名判断文件类型

    Args:
        filename: 文件名

    Returns:
        str: FileType枚举值
    """
    filename = filename.lower()
    if re.match(r".*\.pdf$", filename):
        return FileType.PDF

    if re.match(
        r".*\.(msg|eml|doc|docx|ppt|pptx|yml|xml|htm|json|jsonl|ldjson|csv|txt|ini|xls|xlsx|wps|rtf|hlp|pages|numbers|key|md|mdx|py|js|java|c|cpp|h|php|go|ts|sh|cs|kt|html|sql)$",
        filename
    ):
        return FileType.DOC

    if re.match(r".*\.(wav|flac|ape|alac|wavpack|wv|mp3|aac|ogg|vorbis|opus|m4a|wma|aiff|aif)$", filename):
        return FileType.AURAL

    if re.match(
        r".*\.(jpg|jpeg|png|tif|gif|pcx|tga|exif|fpx|svg|psd|cdr|pcd|dxf|ufo|eps|ai|raw|webp|avif|apng|icon|ico|mpg|mpeg|avi|rm|rmvb|mov|wmv|asf|dat|asx|wvx|mpe|mpa|mp4|mkv|webm|m4v|flv|ts|vob)$",
        filename
    ):
        return FileType.VISUAL

    return FileType.OTHER


def get_file_suffix(filename):
    """
    获取文件扩展名（不含点号）

    Args:
        filename: 文件名

    Returns:
        str: 文件扩展名
    """
    return Path(filename).suffix.lstrip(".").lower()


def get_mime_type(filename):
    """
    根据文件扩展名获取MIME类型

    Args:
        filename: 文件名

    Returns:
        str: MIME类型字符串
    """
    import mimetypes
    mime_type, _ = mimetypes.guess_type(filename)
    return mime_type or "application/octet-stream"


def thumbnail_img(filename, blob):
    """
    生成缩略图二进制数据，控制大小不超过65535字节

    Args:
        filename: 文件名
        blob: 文件二进制数据

    Returns:
        bytes or None: 缩略图二进制数据，不支持则返回None
    """
    file_type = filename_type(filename)

    if file_type == FileType.PDF:
        if not PDFPLUMBER_AVAILABLE:
            return _get_vscode_icon(filename)
        try:
            with sys.modules[LOCK_KEY_pdfplumber]:
                pdf = pdfplumber.open(BytesIO(blob))
                buffered = BytesIO()
                resolution = 32
                img = None
                for _ in range(10):
                    page = pdf.pages[0]
                    page_image = page.to_image(resolution=resolution)
                    page_image.annotated.save(buffered, format="png")
                    img = buffered.getvalue()
                    if len(img) >= 64000 and resolution >= 2:
                        resolution = resolution / 2
                        buffered = BytesIO()
                    else:
                        break
                pdf.close()
                return img
        except Exception:
            return _get_vscode_icon(filename)

    if file_type == FileType.VISUAL:
        if not PIL_AVAILABLE:
            return _get_vscode_icon(filename)
        try:
            image = Image.open(BytesIO(blob))
            image.thumbnail((30, 30))
            buffered = BytesIO()
            image.save(buffered, format="png")
            return buffered.getvalue()
        except Exception:
            return _get_vscode_icon(filename)

    return _get_vscode_icon(filename)


def _get_vscode_icon(filename):
    """
    使用VSCodeIcons库获取文件图标

    Args:
        filename: 文件名

    Returns:
        bytes or None: 图标二进制数据，获取失败则返回None
    """
    try:
        from vscode_icons.vscode import VSCodeIcons
        import importlib.resources
        
        vsi = VSCodeIcons()
        icon_path = vsi.findFileIcon(filename)
        
        if icon_path:
            if 'default' in icon_path.lower():
                extension = filename.split('.')[-1].lower() if '.' in filename else 'txt'
                custom_icon_path = get_file_icon_path(extension)
                if custom_icon_path.exists():
                    with open(custom_icon_path, 'rb') as f:
                        return f.read()
            
            try:
                if not icon_path.startswith('static/'):
                    icon_path = f'static/{icon_path}'
                with importlib.resources.path('vscode_icons', icon_path) as icon_full_path:
                    if icon_full_path.exists():
                        with open(icon_full_path, 'rb') as f:
                            icon_data = f.read()
                        
                        if icon_path.lower().endswith('.svg'):
                            return icon_data
                        elif PIL_AVAILABLE:
                            try:
                                image = Image.open(BytesIO(icon_data))
                                image.thumbnail((32, 32))
                                buffered = BytesIO()
                                image.save(buffered, format="png")
                                return buffered.getvalue()
                            except Exception:
                                return None
                        return None
            except (FileNotFoundError, ModuleNotFoundError):
                pass
    except Exception as e:
        pass
    
    return None


def thumbnail(filename, blob):
    """
    生成base64编码的缩略图字符串

    Args:
        filename: 文件名
        blob: 文件二进制数据

    Returns:
        str: base64编码的缩略图字符串，不支持则返回空字符串
    """
    img = thumbnail_img(filename, blob)
    if img is not None:
        encoded = base64.b64encode(img).decode("utf-8")
        if img.startswith(b'<svg'):
            return "data:image/svg+xml;base64," + encoded
        else:
            return IMG_BASE64_PREFIX + encoded
    return ""


def duplicate_filename(kb_id, filename):
    """
    处理同名文件，在文件名后添加递增数字后缀

    如果文件目录中有相同文件则需要在文件名后面加（递增数字）：
    比如上传文件test.docx,如果存在则需要修改为test_(1).docx,
    如果存在test_(1).docx则修改为test_(2).docx

    Args:
        kb_id: 知识库ID
        filename: 原始文件名

    Returns:
        str: 去重后的文件名
    """
    from app.database.storage.rustfs_utils import rustfs_utils

    if not rustfs_utils.is_available:
        return filename

    bucket_name = kb_id
    if not rustfs_utils.bucket_exists(bucket_name):
        return filename

    base_name, ext = _split_filename(filename)
    object_key = f"{base_name}{ext}" if ext else base_name

    if not rustfs_utils.object_exists(bucket_name, object_key):
        return filename

    counter = 1
    while True:
        new_name = f"{base_name}_({counter}){ext}" if ext else f"{base_name}_({counter})"
        if not rustfs_utils.object_exists(bucket_name, new_name):
            return new_name
        counter += 1


def _split_filename(filename):
    """
    分割文件名为基础名和扩展名

    Args:
        filename: 文件名

    Returns:
        tuple: (base_name, extension) 如 ("test", ".docx")
    """
    path = Path(filename)
    base = path.stem
    ext = path.suffix
    return base, ext


def get_chunk_method_by_file_type(file_type, filename, default_chunk_method="naive"):
    """
    根据文件类型和文件名推断默认的chunk方法

    Args:
        file_type: FileType枚举值
        filename: 文件名
        default_chunk_method: 默认chunk方法

    Returns:
        str: chunk方法名称
    """
    from app.constants.knowledgebase_document_constants import get_default_chunk_method, ChunkMethod
    
    return get_default_chunk_method(file_type, filename)


def find_ffmpeg():
    """查找ffmpeg可执行文件"""
    try:
        import subprocess
        result = subprocess.run(['ffmpeg', '-version'], capture_output=True, timeout=5)
        if result.returncode == 0:
            logger.info("在系统PATH中找到ffmpeg")
            return 'ffmpeg'
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    
    for loc in FFMPEG_COMMON_LOCATIONS:
        if os.path.exists(loc):
            logger.info(f"在常见位置找到ffmpeg: {loc}")
            return loc
    
    logger.warning("未找到ffmpeg")
    return None


def get_ffmpeg_path():
    """获取ffmpeg路径"""
    ffmpeg_path = os.environ.get('FFMPEG_PATH')
    if ffmpeg_path and os.path.exists(ffmpeg_path):
        logger.info(f"从环境变量FFMPEG_PATH获取ffmpeg: {ffmpeg_path}")
        return ffmpeg_path
    return find_ffmpeg()


def cleanup_temp_files(*file_paths):
    """清理临时文件"""
    for file_path in file_paths:
        if file_path and os.path.exists(file_path):
            try:
                os.unlink(file_path)
            except Exception:
                pass


def convert_to_wav(audio_path: str) -> tuple:
    """将音频文件转换为wav格式
    
    Args:
        audio_path: 原始音频文件路径
    
    Returns:
        tuple: (转换后的wav文件路径, 错误信息)，成功时错误信息为None
    """
    if os.path.splitext(audio_path)[1].lower() == '.wav':
        return audio_path, None
    
    ffmpeg_exe = get_ffmpeg_path()
    
    try:
        from pydub import AudioSegment
        if ffmpeg_exe and ffmpeg_exe != 'ffmpeg':
            ffmpeg_dir = os.path.dirname(ffmpeg_exe)
            if ffmpeg_dir not in os.environ.get('PATH', ''):
                os.environ['PATH'] = ffmpeg_dir + os.pathsep + os.environ.get('PATH', '')
                logger.info(f"已将ffmpeg目录添加到PATH: {ffmpeg_dir}")
        
        temp_wav = tempfile.NamedTemporaryFile(suffix='.wav', delete=False)
        temp_wav.close()
        
        audio = AudioSegment.from_file(audio_path)
        audio.export(temp_wav.name, format='wav')
        logger.info(f"成功使用pydub转换音频文件: {audio_path} -> {temp_wav.name}")
        return temp_wav.name, None
    except ImportError:
        logger.warning("pydub未安装，尝试使用ffmpeg")
        if not ffmpeg_exe:
            return None, "音频转换失败: 未找到ffmpeg。请安装ffmpeg或将音频文件转换为wav格式"
        
        try:
            import subprocess
            temp_wav = tempfile.NamedTemporaryFile(suffix='.wav', delete=False)
            temp_wav.close()
            
            subprocess.run([ffmpeg_exe, '-i', audio_path, '-y', temp_wav.name], 
                          check=True, capture_output=True, text=True)
            logger.info(f"成功使用ffmpeg转换音频文件: {audio_path} -> {temp_wav.name}")
            return temp_wav.name, None
        except subprocess.CalledProcessError as e:
            return None, f"音频转换失败: ffmpeg转换失败 - {e.stderr}"
        except Exception as e:
            return None, f"音频转换失败: {str(e)}"
    except Exception as e:
        error_msg = f"音频转换失败: {str(e)}"
        if "ffmpeg" in str(e).lower() or "file not found" in str(e).lower():
            if not ffmpeg_exe:
                error_msg = "音频转换失败: pydub需要ffmpeg支持。请安装ffmpeg或将音频文件转换为wav格式"
        return None, error_msg


def convert_base64_audio_to_wav(base64_content: str, original_filename: str) -> tuple:
    """将base64编码的音频转换为wav格式
    
    Args:
        base64_content: base64编码的音频数据
        original_filename: 原始文件名，用于确定文件扩展名
    
    Returns:
        tuple: (转换后的wav文件的base64编码, 错误信息)，成功时错误信息为None
    """
    temp_file_path = None
    converted_audio_path = None
    
    try:
        # 先将base64保存为临时文件
        binary_data = base64.b64decode(base64_content)
        suffix = os.path.splitext(original_filename)[1].lower() or '.tmp'
        temp_file = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
        temp_file.write(binary_data)
        temp_file.close()
        temp_file_path = temp_file.name
        
        # 转换为wav
        converted_audio_path, error_msg = convert_to_wav(temp_file_path)
        if error_msg:
            return None, error_msg
        
        # 读取转换后的wav文件并编码为base64
        with open(converted_audio_path, 'rb') as f:
            wav_data = f.read()
        wav_base64 = base64.b64encode(wav_data).decode('utf-8')
        
        return wav_base64, None
    finally:
        cleanup_temp_files(temp_file_path)
        if converted_audio_path and converted_audio_path != temp_file_path:
            cleanup_temp_files(converted_audio_path)


def generate_markdown_file(content: str) -> tuple:
    """
    为富文本内容生成临时markdown文件
    
    Args:
        content: 富文本内容（markdown格式）
    
    Returns:
        tuple: (临时文件路径, 二进制数据, 错误信息)，成功时错误信息为None
    """
    temp_file_path = None
    
    try:
        # 创建临时markdown文件
        temp_file = tempfile.NamedTemporaryFile(suffix='.md', delete=False, mode='w', encoding='utf-8')
        temp_file.write(content)
        temp_file.close()
        temp_file_path = temp_file.name
        
        # 读取二进制数据
        with open(temp_file_path, 'rb') as f:
            binary_data = f.read()
        
        logger.info(f"成功生成markdown临时文件: {temp_file_path}")
        return temp_file_path, binary_data, None
    except Exception as e:
        error_msg = f"生成markdown文件失败: {str(e)}"
        logger.error(error_msg)
        if temp_file_path:
            cleanup_temp_files(temp_file_path)
        return None, None, error_msg


def generate_custom_template_excel(document_config: Dict[str, Any]) -> tuple:
    """
    为自定义模版知识生成临时excel文件
    
    Args:
        document_config: 文档配置对象，包含custom_fields和chapters
    
    Returns:
        tuple: (临时文件路径, 二进制数据, 错误信息)，成功时错误信息为None
    """
    temp_file_path = None
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "知识内容"
        
        # 获取自定义字段
        custom_fields = document_config.get('custom_fields', [])
        chapters = document_config.get('chapters', [])
        
        # 构建表头
        headers = []
        for field in custom_fields:
            headers.append(field.get('field_name', ''))
        headers.append('章节')
        
        # 写入表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 构建数据行
        rows_data = []
        
        # 获取富文本内容（用于富文本章节类型）
        content = document_config.get('content', '')
        
        # 如果没有章节但有富文本内容，生成一行（自定义字段值 + 富文本内容）
        if not chapters and content:
            row_values = []
            for field in custom_fields:
                field_value = field.get('value', '')
                row_values.append(str(field_value) if field_value is not None else '')
            row_values.append(content)
            rows_data.append(row_values)
        # 如果没有章节且没有富文本内容，生成一行（自定义字段值 + 空章节内容）
        elif not chapters:
            row_values = []
            for field in custom_fields:
                field_value = field.get('value', '')
                row_values.append(str(field_value) if field_value is not None else '')
            row_values.append('')
            rows_data.append(row_values)
        else:
            # 递归处理章节
            def process_chapters(chapter_list: List[Dict], parent_path: str = ''):
                for chapter in chapter_list:
                    chapter_name = chapter.get('name', '')
                    chapter_path = f"{parent_path}/{chapter_name}" if parent_path else f"/{chapter_name}"
                    chapter_type = chapter.get('type', '')
                    chapter_value = chapter.get('value')
                    
                    # 构建章节内容
                    chapter_content = f"  \n章节名称：{chapter_name}  \n"
                    chapter_content += f"章节路径: {chapter_path}  \n"
                    chapter_content += "章节内容：  \n"
                    
                    if chapter_type == 'form':
                        # 表单类型：字段名：字段值，每个字段换行
                        if chapter_value and isinstance(chapter_value, dict):
                            fields = chapter.get('fields', [])
                            for field in fields:
                                field_id = field.get('id', '')
                                field_name = field.get('field_name', '')
                                field_value = chapter_value.get(field_id, '')
                                chapter_content += f"{field_name}：{field_value}  \n"
                    
                    elif chapter_type == 'list':
                        # 列表类型：组装成表格的html字符串
                        if chapter_value and isinstance(chapter_value, list):
                            fields = chapter.get('fields', [])
                            table_html = "<table>  \n"
                            # 表头
                            table_html += "<tr>"
                            for field in fields:
                                field_name = field.get('field_name', '')
                                table_html += f"<th>{field_name}</th>"
                            table_html += "</tr>  \n"
                            # 表格内容
                            for row_data in chapter_value:
                                table_html += "<tr>"
                                for field in fields:
                                    field_id = field.get('id', '')
                                    cell_value = row_data.get(field_id, '')
                                    table_html += f"<td>{cell_value}</td>"
                                table_html += "</tr>  \n"
                            table_html += "</table>"
                            chapter_content += table_html
                    
                    elif chapter_type == 'rich_text':
                        # 富文本类型：直接使用内容
                        if chapter_value:
                            chapter_content += str(chapter_value)
                    
                    # 构建行数据
                    row_values = []
                    for field in custom_fields:
                        field_value = field.get('value', '')
                        row_values.append(str(field_value) if field_value is not None else '')
                    row_values.append(chapter_content)
                    rows_data.append(row_values)
                    
                    # 递归处理子章节
                    child_chapters = [ch for ch in chapters if ch.get('parentId') == chapter.get('id')]
                    if child_chapters:
                        process_chapters(child_chapters, chapter_path)
            
            # 获取根章节（没有parentId的章节）
            root_chapters = [ch for ch in chapters if not ch.get('parentId')]
            process_chapters(root_chapters)
        
        # 写入数据行
        for row_idx, row_data in enumerate(rows_data, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
        
        # 保存到临时文件
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        temp_file_path = temp_file.name
        
        wb.save(temp_file_path)
        
        # 读取二进制数据
        with open(temp_file_path, 'rb') as f:
            binary_data = f.read()
        
        logger.info(f"成功生成excel临时文件: {temp_file_path}")
        return temp_file_path, binary_data, None
    except ImportError as e:
        error_msg = f"生成excel文件失败: openpyxl库未安装 - {str(e)}"
        logger.error(error_msg)
        return None, None, error_msg
    except Exception as e:
        error_msg = f"生成excel文件失败: {str(e)}"
        logger.error(error_msg)
        if temp_file_path:
            cleanup_temp_files(temp_file_path)
        return None, None, error_msg
