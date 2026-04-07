"""
表格文档切片方法
支持Excel、CSV等表格数据的结构化解析和处理
"""

import copy
import re
import csv
import io
import logging
from collections import Counter
from typing import List, Dict, Any, Optional, Tuple

try:
    import numpy as np
    import pandas as pd
except ImportError:
    np = None
    pd = None

try:
    from xpinyin import Pinyin
except ImportError:
    Pinyin = None

try:
    from dateutil.parser import parse as datetime_parse
except ImportError:
    datetime_parse = None

logger = logging.getLogger(__name__)


class Excel:
    """增强的Excel解析器，支持复杂表头和数据类型推断"""
    
    def __call__(self, fnm, binary=None, from_page=0, to_page=10000000000, callback=None, **kwargs):
        """
        解析Excel文件
        
        Args:
            fnm: 文件名
            binary: 二进制数据
            from_page: 起始行
            to_page: 结束行
            callback: 回调函数
            **kwargs: 额外参数
            
        Returns:
            tuple: (DataFrame列表, 表格图片列表)
        """
        if pd is None:
            raise ImportError("请安装pandas和openpyxl: pip install pandas openpyxl")
        
        try:
            from openpyxl import load_workbook
            from io import BytesIO
        except ImportError:
            raise ImportError("请安装openpyxl: pip install openpyxl")
        
        if binary:
            wb = load_workbook(BytesIO(binary), data_only=True)
        else:
            wb = load_workbook(fnm, data_only=True)
        
        dfs = []
        tbls = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            try:
                rows = list(ws.iter_rows(values_only=True))
            except Exception as e:
                logger.warning(f"跳过工作表 '{sheet_name}'，读取错误: {e}")
                continue
                
            if not rows:
                continue
            
            # 解析表头
            headers, header_rows = self._parse_headers(rows)
            if not headers:
                continue
            
            # 提取数据
            data = []
            rn = 0
            for i, row in enumerate(rows[header_rows:], start=header_rows):
                rn += 1
                if rn - 1 < from_page:
                    continue
                if rn - 1 >= to_page:
                    break
                    
                row_data = self._extract_row_data(row, len(headers))
                if row_data and not self._is_empty_row(row_data):
                    data.append(row_data)
            
            if data:
                df = pd.DataFrame(data, columns=headers)
                dfs.append(df)
        
        callback(0.3, f"提取记录完成: 共 {sum(len(df) for df in dfs)} 条")
        return dfs, tbls
    
    def _parse_headers(self, rows):
        """解析表头"""
        if not rows:
            return [], 0
            
        first_row = rows[0]
        headers = []
        
        for cell in first_row:
            if cell is not None:
                val = str(cell).strip()
                headers.append(val if val else f"Column_{len(headers)+1}")
            else:
                headers.append(f"Column_{len(headers)+1}")
        
        return headers, 1
    
    def _extract_row_data(self, row, expected_cols):
        """提取行数据"""
        row_data = []
        for i in range(expected_cols):
            if i < len(row):
                val = row[i]
                row_data.append(val)
            else:
                row_data.append(None)
        return row_data
    
    def _is_empty_row(self, row_data):
        """检查是否为空行"""
        return all(v is None or str(v).strip() == "" for v in row_data)


def trans_datatime(s):
    """转换日期时间字符串"""
    if datetime_parse is None:
        return s
    try:
        return datetime_parse(s.strip()).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def trans_bool(s):
    """转换布尔值字符串"""
    if re.match(r"(true|yes|是|\*|✓|✔|☑|✅|√)$", str(s).strip(), flags=re.IGNORECASE):
        return "yes"
    if re.match(r"(false|no|否|⍻|×)$", str(s).strip(), flags=re.IGNORECASE):
        return "no"
    return None


def column_data_type(arr):
    """
    推断列的数据类型
    
    Returns:
        tuple: (转换后的数组, 类型名称)
    """
    arr = list(arr)
    counts = {"int": 0, "float": 0, "text": 0, "datetime": 0, "bool": 0}
    trans = {
        t: f for f, t in [
            (int, "int"), (float, "float"), (trans_datatime, "datetime"),
            (trans_bool, "bool"), (str, "text")
        ]
    }
    
    float_flag = False
    for a in arr:
        if a is None:
            continue
        if re.match(r"[+-]?[0-9]+$", str(a).replace("%", "")) and not str(a).replace("%", "").startswith("0"):
            counts["int"] += 1
            if int(str(a)) > 2 ** 63 - 1:
                float_flag = True
                break
        elif re.match(r"[+-]?[0-9.]{,19}$", str(a).replace("%", "")) and not str(a).replace("%", "").startswith("0"):
            counts["float"] += 1
        elif re.match(r"(true|yes|是|\*|✓|✔|☑|✅|√|false|no|否|⍻|×)$", str(a), flags=re.IGNORECASE):
            counts["bool"] += 1
        elif trans_datatime(str(a)):
            counts["datetime"] += 1
        else:
            counts["text"] += 1
    
    if float_flag:
        ty = "float"
    else:
        counts = sorted(counts.items(), key=lambda x: x[1] * -1)
        ty = counts[0][0]
    
    for i in range(len(arr)):
        if arr[i] is None:
            continue
        try:
            arr[i] = trans[ty](str(arr[i]))
        except Exception as e:
            arr[i] = None
            logger.warning(f"列 {i}: {e}")
    
    return arr, ty


def chunk(filename, binary=None, from_page=0, to_page=10000000000, lang="Chinese", callback=None, **kwargs):
    """
    表格切片函数
    
    支持Excel、CSV、TXT格式的表格数据解析，每一行作为一个chunk
    
    Args:
        filename: 文件名或文件路径
        binary: 文件二进制数据（可选）
        from_page: 起始行号
        to_page: 结束行号
        lang: 语言 ("Chinese" 或 "English")
        callback: 进度回调函数 callback(progress, message)
        **kwargs: 其他参数，包括:
            - delimiter: CSV/TXT分隔符（默认\t或,）
            
    Returns:
        list: 切片后的文档列表
    """
    from ..nlp import rag_tokenizer, tokenize
    
    if pd is None:
        raise ImportError("请安装pandas: pip install pandas")
    
    is_english = lang.lower() == "english"
    tbls = []
    res = []
    
    # Excel文件处理
    if re.search(r"\.xlsx?$", filename, re.IGNORECASE):
        callback(0.1, "开始解析Excel表格")
        
        excel_parser = Excel()
        dfs, tbls = excel_parser(
            filename, binary, from_page=from_page, to_page=to_page, callback=callback, **kwargs
        )
        
    # TXT文件处理
    elif re.search(r"\.txt$", filename, re.IGNORECASE):
        callback(0.1, "开始解析TXT表格")
        
        txt = _get_text(filename, binary)
        lines = txt.split("\n")
        delimiter = kwargs.get("delimiter", "\t")
        
        headers = lines[0].split(delimiter)
        rows = []
        fails = []
        
        for i, line in enumerate(lines[1:]):
            if i < from_page:
                continue
            if i >= to_page:
                break
            row = [field for field in line.split(delimiter)]
            if len(row) != len(headers):
                fails.append(str(i))
                continue
            rows.append(row)
        
        callback(0.3, f"提取记录: {from_page+1}~{min(to_page, from_page+len(rows))}" + (
            f", {len(fails)}失败, 行: {','.join(fails[:3])}" if fails else ""
        ))
        
        if np is not None:
            dfs = [pd.DataFrame(np.array(rows), columns=headers)]
        else:
            dfs = [pd.DataFrame(rows, columns=headers)]
        
    # CSV文件处理
    elif re.search(r"\.csv$", filename, re.IGNORECASE):
        callback(0.1, "开始解析CSV表格")
        
        txt = _get_text(filename, binary)
        delimiter = kwargs.get("delimiter", ",")
        
        reader = csv.reader(io.StringIO(txt), delimiter=delimiter)
        all_rows = list(reader)
        if not all_rows:
            raise ValueError("空CSV文件")
        
        headers = all_rows[0]
        fails = []
        rows = []
        
        for i, row in enumerate(all_rows[1 + from_page: 1 + to_page]):
            if len(row) != len(headers):
                fails.append(str(i + from_page))
                continue
            rows.append(row)
        
        callback(0.3, f"提取记录: {from_page}~{from_page + len(rows)}" + (
            f", {len(fails)}失败, 行: {','.join(fails[:3])}" if fails else ""
        ))
        
        dfs = [pd.DataFrame(rows, columns=headers)]
        
    else:
        raise NotImplementedError(f"不支持的文件类型: {filename} (支持: xlsx, txt, csv)")
    
    # 处理DataFrame
    PY = Pinyin() if Pinyin else None
    fields_map = {
        "text": "_tks", "int": "_long", "keyword": "_kwd",
        "float": "_flt", "datetime": "_dt", "bool": "_kwd"
    }
    
    for df in dfs:
        # 移除ID列
        for n in ["id", "_id", "index", "idx"]:
            if n in df.columns:
                del df[n]
        
        clmns = df.columns.values
        if len(clmns) != len(set(clmns)):
            col_counts = Counter(clmns)
            duplicates = [col for col, count in col_counts.items() if count > 1]
            raise ValueError(f"检测到重复列名: {duplicates}\n来自: {clmns}")
        
        txts = list(copy.deepcopy(clmns))
        
        # 生成拼音列名（如果可用）
        if PY:
            py_clmns = [
                PY.get_pinyins(re.sub(r"(/.*|（[^（）]+?）|\([^()]+?\))", "", str(n)), "_")[0]
                for n in clmns
            ]
        else:
            py_clmns = [f"col_{i}" for i in range(len(clmns))]
        
        clmn_tys = []
        for j in range(len(clmns)):
            cln, ty = column_data_type(df[clmns[j]])
            clmn_tys.append(ty)
            df[clmns[j]] = cln
            if ty == "text":
                txts.extend([str(c) for c in cln if c])
        
        clmns_map = [
            (py_clmns[i].lower() + fields_map[clmn_tys[i]], str(clmns[i]).replace("_", " "))
            for i in range(len(clmns))
        ]
        
        eng = lang.lower() == "english"
        
        for ii, row in df.iterrows():
            d = {"docnm_kwd": filename, "title_tks": rag_tokenizer.tokenize(re.sub(r"\.[a-zA-Z]+$", "", filename))}
            row_fields = []
            
            for j in range(len(clmns)):
                if row[clmns[j]] is None:
                    continue
                if not str(row[clmns[j]]):
                    continue
                if not isinstance(row[clmns[j]], pd.Series) and pd.isna(row[clmns[j]]):
                    continue
                    
                fld = clmns_map[j][0]
                d[fld] = row[clmns[j]] if clmn_tys[j] != "text" else rag_tokenizer.tokenize(row[clmns[j]])
                row_fields.append((clmns[j], row[clmns[j]]))
            
            if not row_fields:
                continue
            
            formatted_text = "\n".join([f"- {field}: {value}" for field, value in row_fields])
            tokenize(d, formatted_text, eng)
            res.append(d)
    
    if tbls:
        doc = {"docnm_kwd": filename, "title_tks": rag_tokenizer.tokenize(re.sub(r"\.[a-zA-Z]+$", "", filename))}
        from ..nlp import tokenize_table
        res.extend(tokenize_table(tbls, doc, is_english))
    
    callback(0.35, "")
    return res


def _get_text(filename, binary=None):
    """获取文本内容"""
    from ..nlp import find_codec
    
    if binary:
        encoding = find_codec(binary)
        return binary.decode(encoding, errors="ignore")
    else:
        with open(filename, 'rb') as f:
            blob = f.read()
        encoding = find_codec(blob)
        return blob.decode(encoding, errors="ignore")


if __name__ == "__main__":
    import sys
    
    def dummy_callback(prog=None, msg=""):
        print(f"[{prog}] {msg}" if prog else msg)
    
    if len(sys.argv) > 1:
        result = chunk(sys.argv[1], callback=dummy_callback)
        print(f"Total chunks: {len(result)}")
